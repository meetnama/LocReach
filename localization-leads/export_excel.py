"""
export_excel.py — Export LocReach records to formatted Excel workbooks.

Builders:
  build_excel_bytes(db_path)          — full DB, sheets by domain status + people/leads/blocked
  build_step1_excel_bytes(...)        — current Step 1 run only
  build_people_excel_bytes(...)       — current Step 2 run only
  build_leads_excel_bytes(...)        — current Step 3 run only

Usage (standalone):
    python export_excel.py                  → saves  exports/locreach_YYYYMMDD_HHMM.xlsx
    python export_excel.py myfile.xlsx      → saves  myfile.xlsx
"""

import io
import os
import sys
import sqlite3
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter


# ── Colour palette ─────────────────────────────────────────────────────────────
C_HEADER_BG   = "1E2A3A"   # dark navy   — header fill
C_HEADER_FG   = "BFDBFE"   # light blue  — header text
C_TITLE_BG    = "0F1724"   # near-black  — sheet title row
C_TITLE_FG    = "E8F4F8"   # near-white  — title text
C_ROW_ODD     = "F4F8FB"   # very light blue-grey
C_ROW_EVEN    = "FFFFFF"   # white
C_ACCENT      = "10B981"   # emerald — positive stats
C_ACCENT2     = "3B82F6"   # blue    — neutral stats
C_BORDER      = "D0DDE8"   # subtle border
C_SHEET_TAB   = "1E2A3A"   # (unused — openpyxl can't set tab colour reliably)


# ── Style helpers ──────────────────────────────────────────────────────────────

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(bold=False, color="1A1A2E", size=10, italic=False) -> Font:
    return Font(bold=bold, color=color, size=size, italic=italic,
                name="Calibri")


def _border() -> Border:
    thin = Side(style="thin", color=C_BORDER)
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def _left() -> Alignment:
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


def _apply_header_row(ws, row: int, columns: list[str]) -> None:
    """Write and style a header row."""
    for col_idx, label in enumerate(columns, 1):
        cell = ws.cell(row=row, column=col_idx, value=label)
        cell.fill      = _fill(C_HEADER_BG)
        cell.font      = _font(bold=True, color=C_HEADER_FG, size=10)
        cell.alignment = _center()
        cell.border    = _border()


def _apply_data_row(ws, row: int, values: list, odd: bool) -> None:
    """Write and style a data row."""
    bg = C_ROW_ODD if odd else C_ROW_EVEN
    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col_idx, value=val if val is not None else "")
        cell.fill      = _fill(bg)
        cell.font      = _font(size=9)
        cell.alignment = _left()
        cell.border    = _border()


def _auto_col_widths(ws, min_w=8, max_w=50) -> None:
    """Set column widths based on content."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max(min_w, min(max_len + 3, max_w))


def _title_row(ws, row: int, text: str, ncols: int) -> None:
    """Full-width title row at the top of a sheet."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    cell = ws.cell(row=row, column=1, value=text)
    cell.fill      = _fill(C_TITLE_BG)
    cell.font      = _font(bold=True, color=C_TITLE_FG, size=12)
    cell.alignment = _center()
    ws.row_dimensions[row].height = 28


def _freeze(ws, row: int, col: int) -> None:
    from openpyxl.utils.cell import get_column_letter
    ws.freeze_panes = f"{get_column_letter(col)}{row}"


# ── Sheet builders ─────────────────────────────────────────────────────────────

def _sheet_summary(wb: Workbook, stats: dict) -> None:
    ws = wb.create_sheet("Summary", 0)
    ws.sheet_view.showGridLines = False

    _title_row(ws, 1, "LocReach Lead Discovery — Full Database Export", 3)

    ws.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d  %H:%M')}")
    ws.cell(row=2, column=1).font      = _font(italic=True, color="888888", size=9)
    ws.cell(row=2, column=1).alignment = _left()
    ws.row_dimensions[2].height = 18

    # Stat cards (key / value / note) — domains split by status category
    by_status = stats.get("by_status") or {}
    data = [
        ("Qualified Domains",   by_status.get("qualified", 0),   "Passed score + geo gate"),
        ("Rejected Domains",    by_status.get("rejected", 0),    "Scored but not kept (industry/geo/tier)"),
        ("Failed Domains",      by_status.get("failed", 0),      "Scrape/score errors"),
        ("Unreachable Domains", by_status.get("unreachable", 0), "Site could not be reached"),
        ("Discovered Domains",  by_status.get("discovered", 0),  "Found but not yet processed"),
        ("People",              stats.get("people", 0),          "Step 2 contacts"),
        ("Leads with Emails",   stats.get("leads", 0),           "Step 3 confirmed emails"),
        ("Blocked Domains",     stats.get("blocked", 0),         "Junk / directory sites filtered out"),
    ]

    headers = ["Metric", "Count", "Description"]
    _apply_header_row(ws, 4, headers)

    for i, (metric, count, note) in enumerate(data, 5):
        _apply_data_row(ws, i, [metric, count, note], odd=(i % 2 == 1))
        c = ws.cell(row=i, column=2)
        c.font = Font(bold=True, color=C_ACCENT if count > 0 else "888888",
                      size=11, name="Calibri")
        c.alignment = _center()

    # Industry breakdown
    if stats.get("by_industry"):
        row = len(data) + 7
        ws.cell(row=row, column=1, value="Breakdown by Industry").font = \
            _font(bold=True, color=C_HEADER_BG, size=10)
        row += 1
        _apply_header_row(ws, row, ["Industry", "Domains Found"])
        row += 1
        for j, (ind, cnt) in enumerate(stats["by_industry"]):
            _apply_data_row(ws, row + j, [ind or "(all)", cnt], odd=(j % 2 == 0))

    # Country breakdown
    if stats.get("by_country"):
        row2 = len(data) + 7
        col_offset = 4
        ws.cell(row=row2, column=col_offset, value="Breakdown by Country").font = \
            _font(bold=True, color=C_HEADER_BG, size=10)
        row2 += 1
        for col_idx, label in enumerate(["Country", "Domains Found"], col_offset):
            cell = ws.cell(row=row2, column=col_idx, value=label)
            cell.fill      = _fill(C_HEADER_BG)
            cell.font      = _font(bold=True, color=C_HEADER_FG, size=10)
            cell.alignment = _center()
            cell.border    = _border()
        row2 += 1
        for j, (cntry, cnt) in enumerate(stats["by_country"]):
            for ci, val in enumerate([cntry or "(all)", cnt], col_offset):
                cell = ws.cell(row=row2 + j, column=ci, value=val)
                cell.fill      = _fill(C_ROW_ODD if j % 2 == 0 else C_ROW_EVEN)
                cell.font      = _font(size=9)
                cell.alignment = _left()
                cell.border    = _border()

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 45
    ws.column_dimensions["D"].width = 24
    ws.column_dimensions["E"].width = 10


_DOMAIN_SHEET_COLS = [
    "#", "Domain", "Company", "Industry", "Country", "Keyword",
    "Score", "Tier", "Signals", "LinkedIn", "Found At", "Qualified At",
]


def _fmt_ts(val) -> str:
    if not val:
        return "—"
    return str(val)[:16].replace("T", "  ")


def _sheet_domains_category(wb: Workbook, title: str, rows: list) -> None:
    """One sheet for a single domain status category (qualified, rejected, …)."""
    # Excel sheet names max 31 chars
    safe_title = title[:31]
    ws = wb.create_sheet(safe_title)
    ws.sheet_view.showGridLines = False

    cols = _DOMAIN_SHEET_COLS
    _title_row(ws, 1, f"{title}  ({len(rows)} total)", len(cols))
    _apply_header_row(ws, 2, cols)
    _freeze(ws, 3, 2)

    for i, row in enumerate(rows, 1):
        (domain, company, industry, country, keyword, score, tier,
         reasons, linkedin, found_at, qualified_at) = row
        vals = [
            i, domain, company or "—", industry or "—", country or "—",
            keyword or "—", score if score is not None else 0, tier or "—",
            reasons or "—", linkedin or "—",
            _fmt_ts(found_at), _fmt_ts(qualified_at),
        ]
        _apply_data_row(ws, i + 2, vals, odd=(i % 2 == 1))
        if linkedin:
            cell = ws.cell(row=i + 2, column=10)
            cell.hyperlink = linkedin
            cell.font = Font(color="2563EB", underline="single",
                             size=9, name="Calibri")

    _auto_col_widths(ws)


def _sheet_people_db(wb: Workbook, rows: list) -> None:
    ws = wb.create_sheet("People")
    ws.sheet_view.showGridLines = False

    cols = ["#", "Full Name", "Title", "Company", "Domain", "Type", "LinkedIn"]
    _title_row(ws, 1, f"People  ({len(rows)} total)", len(cols))
    _apply_header_row(ws, 2, cols)
    _freeze(ws, 3, 2)

    for i, row in enumerate(rows, 1):
        full_name, title, company, domain, company_type, linkedin = row
        vals = [
            i, full_name or "—", title or "—", company or "—",
            domain or "—", (company_type or "").upper() or "—", linkedin or "—",
        ]
        _apply_data_row(ws, i + 2, vals, odd=(i % 2 == 1))
        if linkedin:
            cell = ws.cell(row=i + 2, column=7)
            cell.hyperlink = linkedin
            cell.font = Font(color="2563EB", underline="single",
                             size=9, name="Calibri")

    _auto_col_widths(ws)


def _sheet_leads(wb: Workbook, rows: list) -> None:
    ws = wb.create_sheet("Leads")
    ws.sheet_view.showGridLines = False

    cols = ["#", "Full Name", "Title", "Company", "Domain", "Country",
            "Email", "Email Source", "LinkedIn"]
    _title_row(ws, 1, f"Leads — People & Emails  ({len(rows)} total)", len(cols))
    _apply_header_row(ws, 2, cols)
    _freeze(ws, 3, 2)

    for i, row in enumerate(rows, 1):
        email, source, full_name, first, last, title, company, domain, country, linkedin = row
        name = full_name or f"{first or ''} {last or ''}".strip() or "—"
        vals = [i, name, title or "—", company or "—", domain or "—",
                country or "—", email, source or "—", linkedin or "—"]
        _apply_data_row(ws, i + 2, vals, odd=(i % 2 == 1))
        if linkedin:
            cell = ws.cell(row=i + 2, column=9)
            cell.hyperlink = linkedin
            cell.font = Font(color="2563EB", underline="single",
                             size=9, name="Calibri")

    _auto_col_widths(ws)


def _sheet_blocked(wb: Workbook, rows: list) -> None:
    ws = wb.create_sheet("Blocked")
    ws.sheet_view.showGridLines = False

    cols = ["#", "Domain", "Blocked At"]
    _title_row(ws, 1, f"Blocked Domains  ({len(rows)} total)", len(cols))
    _apply_header_row(ws, 2, cols)
    _freeze(ws, 3, 2)

    for i, row in enumerate(rows, 1):
        domain, blocked_at = row
        _apply_data_row(
            ws, i + 2,
            [i, domain, _fmt_ts(blocked_at)],
            odd=(i % 2 == 1),
        )

    _auto_col_widths(ws)


# ── Main builder ───────────────────────────────────────────────────────────────

def build_step1_excel_bytes(
    qualified_rows: list[dict],
    rejected_rows: list[dict] | None = None,
) -> bytes:
    """
    Export Step 1 results from the current run only (not full DB).
    Qualified sheet always; optional Not Kept sheet from rejected_log.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Qualified"
    ws.sheet_view.showGridLines = False

    cols = ["Domain", "Company", "Tier", "Score", "LinkedIn", "Country", "Signals"]
    _title_row(ws, 1, "LocReach Step 1 — Qualified Domains (this run)", len(cols))
    ws.cell(row=2, column=1,
            value=f"Generated: {datetime.now().strftime('%Y-%m-%d  %H:%M')}  ·  {len(qualified_rows)} companies")
    ws.cell(row=2, column=1).font = _font(italic=True, color="888888", size=9)

    _apply_header_row(ws, 4, cols)
    for i, entry in enumerate(
        sorted(qualified_rows, key=lambda e: e.get("score", 0), reverse=True),
        start=5,
    ):
        reasons = entry.get("reasons") or []
        if isinstance(reasons, list):
            reasons = ", ".join(reasons)
        _apply_data_row(
            ws, i,
            [
                entry.get("domain", ""),
                entry.get("company_name", ""),
                entry.get("tier", ""),
                entry.get("score", 0),
                entry.get("linkedin_url", ""),
                entry.get("country", ""),
                reasons,
            ],
            odd=(i % 2 == 1),
        )

    _auto_col_widths(ws)
    _freeze(ws, 5, 1)

    rejected_rows = rejected_rows or []
    if rejected_rows:
        ws2 = wb.create_sheet("Not Kept")
        ws2.sheet_view.showGridLines = False
        cols2 = ["Domain", "Reason", "Detail"]
        _title_row(ws2, 1, f"LocReach Step 1 — Not Kept (this run · {len(rejected_rows)})", len(cols2))
        ws2.cell(row=2, column=1,
                 value=f"Generated: {datetime.now().strftime('%Y-%m-%d  %H:%M')}")
        ws2.cell(row=2, column=1).font = _font(italic=True, color="888888", size=9)
        _apply_header_row(ws2, 4, cols2)
        for i, entry in enumerate(rejected_rows, start=5):
            _apply_data_row(
                ws2, i,
                [
                    entry.get("domain", "") or "",
                    entry.get("reason", "") or "",
                    entry.get("detail", "") or entry.get("message", "") or "",
                ],
                odd=(i % 2 == 1),
            )
        _auto_col_widths(ws2)
        _freeze(ws2, 5, 1)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_people_excel_bytes(people: list[dict]) -> bytes:
    """
    Export the CURRENT People results (this run only) to a formatted workbook.
    `people` are the session result dicts shown on the Step 2 page.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "People"
    ws.sheet_view.showGridLines = False

    cols = ["#", "Full Name", "Title", "Company", "Domain", "Type", "LinkedIn"]
    _title_row(ws, 1, f"LocReach — People (this run · {len(people)} total)", len(cols))
    ws.cell(row=2, column=1,
            value=f"Generated: {datetime.now().strftime('%Y-%m-%d  %H:%M')}")
    ws.cell(row=2, column=1).font = _font(italic=True, color="888888", size=9)

    _apply_header_row(ws, 4, cols)
    for i, p in enumerate(people, start=5):
        _apply_data_row(
            ws, i,
            [
                i - 4,
                p.get("full_name", "") or "",
                p.get("title", "") or "",
                p.get("company_name", "") or "",
                p.get("domain", "") or "",
                (p.get("company_type") or "").upper() or "—",
                p.get("linkedin_url", "") or "",
            ],
            odd=(i % 2 == 1),
        )

    _auto_col_widths(ws)
    _freeze(ws, 5, 1)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_leads_excel_bytes(leads: list[dict]) -> bytes:
    """
    Export the CURRENT Emails/Leads results (this run only) to a formatted
    workbook. `leads` are the session result dicts shown on the Step 3 page.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"
    ws.sheet_view.showGridLines = False

    cols = ["#", "Full Name", "Title", "Company", "Email", "Email Source",
            "Domain", "LinkedIn"]
    _title_row(ws, 1, f"LocReach — Leads (this run · {len(leads)} total)", len(cols))
    ws.cell(row=2, column=1,
            value=f"Generated: {datetime.now().strftime('%Y-%m-%d  %H:%M')}")
    ws.cell(row=2, column=1).font = _font(italic=True, color="888888", size=9)

    _apply_header_row(ws, 4, cols)
    for i, ld in enumerate(leads, start=5):
        _apply_data_row(
            ws, i,
            [
                i - 4,
                ld.get("full_name", "") or "",
                ld.get("title", "") or "",
                ld.get("company", "") or "",
                ld.get("email", "") or "",
                ld.get("email_source", "") or "",
                ld.get("domain", "") or "",
                ld.get("linkedin_url", "") or "",
            ],
            odd=(i % 2 == 1),
        )

    _auto_col_widths(ws)
    _freeze(ws, 5, 1)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_excel_bytes(db_path: str) -> bytes:
    """
    Build a formatted Excel workbook from the full DB.
    Domains are split into separate sheets by status category
    (Qualified, Rejected, Failed, Unreachable, Discovered), plus
    People, Leads, Blocked, and Summary.
    """
    from db import db_init

    conn = sqlite3.connect(db_path)
    db_init(conn)

    domain_sql = (
        "SELECT domain, company_name, industry, country, keyword, "
        "       score, score_tier, score_reasons, linkedin_url, "
        "       found_at, qualified_at "
        "FROM domains WHERE status=? "
        "ORDER BY score DESC, id DESC"
    )
    status_order = (
        ("qualified", "Qualified"),
        ("rejected", "Rejected"),
        ("failed", "Failed"),
        ("unreachable", "Unreachable"),
        ("discovered", "Discovered"),
    )
    domains_by_status = {
        status: conn.execute(domain_sql, (status,)).fetchall()
        for status, _title in status_order
    }

    people = conn.execute(
        "SELECT p.full_name, p.title, p.company_name, p.domain, "
        "       IFNULL(d.company_type, ''), p.linkedin_url "
        "FROM people p "
        "LEFT JOIN domains d ON d.domain = p.domain "
        "ORDER BY p.id DESC"
    ).fetchall()

    leads = conn.execute(
        "SELECT email, email_source, full_name, first_name, last_name, "
        "       title, company, domain, country, linkedin_url "
        "FROM leads ORDER BY id DESC"
    ).fetchall()

    blocked = conn.execute(
        "SELECT domain, blocked_at FROM blocked_domains ORDER BY blocked_at DESC"
    ).fetchall()

    by_industry = conn.execute(
        "SELECT industry, COUNT(*) AS n FROM domains "
        "GROUP BY industry ORDER BY n DESC"
    ).fetchall()

    by_country = conn.execute(
        "SELECT country, COUNT(*) AS n FROM domains "
        "GROUP BY country ORDER BY n DESC"
    ).fetchall()

    conn.close()

    by_status_counts = {
        status: len(rows) for status, rows in domains_by_status.items()
    }
    stats = {
        "by_status":   by_status_counts,
        "people":      len(people),
        "leads":       len(leads),
        "blocked":     len(blocked),
        "by_industry": by_industry,
        "by_country":  by_country,
    }

    wb = Workbook()
    wb.remove(wb.active)

    _sheet_summary(wb, stats)
    for status, title in status_order:
        rows = domains_by_status[status]
        if rows or status == "qualified":
            _sheet_domains_category(wb, title, rows)
    _sheet_people_db(wb, people)
    _sheet_leads(wb, leads)
    _sheet_blocked(wb, blocked)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── CLI entry point ────────────────────────────────────────────────────────────

if __name__ == "__main__":
    base = os.path.dirname(os.path.abspath(__file__))
    db   = os.path.join(base, "leads.db")

    if len(sys.argv) > 1:
        out_path = sys.argv[1]
    else:
        os.makedirs(os.path.join(base, "exports"), exist_ok=True)
        out_path = os.path.join(
            base, "exports",
            f"locreach_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        )

    data = build_excel_bytes(db)
    with open(out_path, "wb") as f:
        f.write(data)
    print(f"Saved: {out_path}")
